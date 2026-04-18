"""Raporlama ve Analiz API Endpoints"""

import json
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from statistics import median
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from api.core.authz import is_global_procurement_manager
from api.database import get_db
from api.models import (
    Quote,
    QuoteItem,
    SupplierQuote,
    SupplierQuoteItem,
    Supplier,
    User,
)
from api.models.report import SupplierRating, PriceAnalysis, Contract
from api.schemas.report import (
    SupplierRatingCreate,
    SupplierRatingOut,
    PriceAnalysisOut,
    QuoteComparisonOut,
    QuoteComparisonMetric,
    ReportingDashboardData,
    ContractOut,
)
from api.core.deps import get_current_user
from api.core.time import utcnow

router = APIRouter(prefix="/reports", tags=["reports"])


def _can_access_quote(current_user: User, quote: Quote) -> bool:
    if is_global_procurement_manager(current_user):
        return True
    if quote.created_by_id == current_user.id:
        return True
    return any(project.id == quote.project_id for project in current_user.projects)


def _get_scoped_quote_or_404(
    db: Session,
    quote_id: int,
    current_user: User,
) -> Quote:
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if quote is None:
        raise HTTPException(status_code=404, detail="Teklif bulunamadı")
    if not _can_access_quote(current_user, quote):
        raise HTTPException(
            status_code=403, detail="Bu raporu görüntüleme yetkiniz yok"
        )
    return quote


def _parse_item_meta(notes: str | None) -> tuple[str, str]:
    if not notes:
        return "", ""
    try:
        parsed = json.loads(notes)
        if isinstance(parsed, dict):
            detail = str(parsed.get("detail") or "")
            image_url = str(parsed.get("image_url") or "")
            return detail, image_url
    except Exception:
        pass
    return str(notes), ""


def _pick_latest_quote_by_supplier(
    supplier_quotes: list[SupplierQuote],
) -> list[SupplierQuote]:
    latest_by_supplier: dict[int, SupplierQuote] = {}

    for sq in supplier_quotes:
        current = latest_by_supplier.get(sq.supplier_id)
        if current is None:
            latest_by_supplier[sq.supplier_id] = sq
            continue

        sq_revision = int(sq.revision_number or 0)
        cur_revision = int(current.revision_number or 0)

        if sq_revision > cur_revision:
            latest_by_supplier[sq.supplier_id] = sq
            continue

        if sq_revision == cur_revision:
            sq_ts = sq.submitted_at or sq.updated_at or sq.created_at
            cur_ts = current.submitted_at or current.updated_at or current.created_at
            if sq_ts and cur_ts and sq_ts > cur_ts:
                latest_by_supplier[sq.supplier_id] = sq

    return sorted(
        latest_by_supplier.values(),
        key=lambda row: float(row.final_amount or 0),
    )


def _build_comparison_dataset(db: Session, quote_id: int) -> dict:
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="Teklif bulunamadı")

    all_supplier_quotes = (
        db.query(SupplierQuote).filter(SupplierQuote.quote_id == quote_id).all()
    )
    if not all_supplier_quotes:
        raise HTTPException(
            status_code=404, detail="Rapor için tedarikçi teklifi bulunamadı"
        )

    latest_quotes = _pick_latest_quote_by_supplier(all_supplier_quotes)
    latest_ids = [row.id for row in latest_quotes]

    supplier_item_rows = (
        db.query(SupplierQuoteItem)
        .filter(SupplierQuoteItem.supplier_quote_id.in_(latest_ids or [-1]))
        .all()
    )
    supplier_item_map = {
        (row.supplier_quote_id, row.quote_item_id): row for row in supplier_item_rows
    }

    quote_items = (
        db.query(QuoteItem)
        .filter(QuoteItem.quote_id == quote_id)
        .order_by(QuoteItem.sequence.asc(), QuoteItem.id.asc())
        .all()
    )

    approved_row = next(
        (row for row in latest_quotes if row.status == "onaylandı"), None
    )
    approved_supplier_name = (
        approved_row.supplier.company_name
        if approved_row and approved_row.supplier
        else ""
    )

    suppliers = []
    for row in latest_quotes:
        suppliers.append(
            {
                "supplier_quote_id": row.id,
                "supplier_id": row.supplier_id,
                "supplier_name": row.supplier.company_name
                if row.supplier
                else f"Supplier#{row.supplier_id}",
                "revision_number": int(row.revision_number or 0),
                "status": row.status,
                "total_amount": float(row.total_amount or 0),
                "discount_amount": float(row.discount_amount or 0),
                "final_amount": float(row.final_amount or 0),
                "delivery_time": int(row.delivery_time or 0),
                "approved": row.status == "onaylandı",
            }
        )

    items = []
    for qi in quote_items:
        detail, image_url = _parse_item_meta(qi.notes)
        item_entry = {
            "quote_item_id": qi.id,
            "line_number": qi.line_number,
            "description": qi.description,
            "detail": detail,
            "image_url": image_url,
            "unit": qi.unit,
            "quantity": float(qi.quantity or 0),
            "base_unit_price": float(qi.unit_price or 0),
            "is_group_header": bool(qi.is_group_header),
            "supplier_prices": {},
        }

        for supplier in suppliers:
            sqi = supplier_item_map.get((supplier["supplier_quote_id"], qi.id))
            item_entry["supplier_prices"][str(supplier["supplier_quote_id"])] = {
                "unit_price": float(sqi.unit_price or 0) if sqi else None,
                "total_price": float(sqi.total_price or 0) if sqi else None,
            }

        items.append(item_entry)

    return {
        "quote": {
            "id": quote.id,
            "rfq_id": quote.id,
            "title": quote.title,
            "generated_at": utcnow().isoformat(),
            "approved_supplier_name": approved_supplier_name,
        },
        "suppliers": suppliers,
        "items": items,
    }


# ============ PRICE ANALYSIS ============


@router.get("/{quote_id}/price-analysis", response_model=PriceAnalysisOut)
def get_price_analysis(
    quote_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Quote için fiyat analiz raporu"""
    quote = _get_scoped_quote_or_404(db, quote_id, current_user)

    # SupplierQuote'ların final_amount'larını al
    supplier_quotes = (
        db.query(SupplierQuote)
        .filter(
            SupplierQuote.quote_id == quote_id,
            SupplierQuote.status.in_(["submitted", "yanıtlandı"]),
        )
        .all()
    )

    if not supplier_quotes:
        raise HTTPException(status_code=404, detail="Yanıt bulunamadı")

    # Fiyatları topla
    prices = [float(sq.final_amount) for sq in supplier_quotes if sq.final_amount]

    if not prices:
        raise HTTPException(status_code=400, detail="Fiyat verisi yok")

    # İstatistikler hesapla
    min_price = float(min(prices))
    max_price = float(max(prices))
    avg_price = float(sum(prices) / len(prices))
    median_price = float(median(prices))
    price_variance = (
        ((max_price - min_price) / min_price * 100) if min_price > 0 else 0.0
    )

    # En ucuz ve en pahalı tedarikçileri bul
    cheapest_sq = sorted(supplier_quotes, key=lambda x: float(x.final_amount))[0]
    most_expensive_sq = sorted(supplier_quotes, key=lambda x: float(x.final_amount))[-1]

    # Raporu kaydet veya güncelle
    analysis = (
        db.query(PriceAnalysis).filter(PriceAnalysis.quote_id == quote_id).first()
    )
    if not analysis:
        analysis = PriceAnalysis(
            quote_id=quote_id,
            min_price=min_price,
            max_price=max_price,
            avg_price=avg_price,
            median_price=median_price,
            price_variance=float(price_variance),
            cheapest_supplier_id=cheapest_sq.supplier_id,
            most_expensive_supplier_id=most_expensive_sq.supplier_id,
            total_responses=len(supplier_quotes),
            submitted_responses=len([sq for sq in supplier_quotes if sq.submitted_at]),
        )
        db.add(analysis)
    else:
        analysis.min_price = min_price  # type: ignore[assignment]
        analysis.max_price = max_price  # type: ignore[assignment]
        analysis.avg_price = avg_price  # type: ignore[assignment]
        analysis.median_price = median_price  # type: ignore[assignment]
        analysis.price_variance = float(price_variance)
        analysis.cheapest_supplier_id = cheapest_sq.supplier_id
        analysis.most_expensive_supplier_id = most_expensive_sq.supplier_id

    db.commit()
    db.refresh(analysis)
    return analysis


# ============ QUOTE COMPARISON ============


@router.get("/{quote_id}/comparison", response_model=QuoteComparisonOut)
def get_quote_comparison(
    quote_id: int,
    comparison_type: str = "price",  # price, delivery, overall
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Quote'un farklı kriterlerine göre karşılaştırması"""
    _get_scoped_quote_or_404(db, quote_id, current_user)

    supplier_quotes = (
        db.query(SupplierQuote)
        .filter(
            SupplierQuote.quote_id == quote_id, SupplierQuote.submitted_at.is_not(None)
        )
        .all()
    )

    if not supplier_quotes:
        raise HTTPException(status_code=404, detail="Gönderilen yanıt yok")

    metrics = []

    if comparison_type == "price":
        # Fiyat karşılaştırması
        prices = [(sq, float(sq.final_amount)) for sq in supplier_quotes]
        prices_sorted = sorted(prices, key=lambda x: x[1])

        min_price = prices_sorted[0][1]
        max_price = prices_sorted[-1][1]
        avg_price = sum(p[1] for p in prices) / len(prices)

        metrics.append(
            QuoteComparisonMetric(
                metric_name="min_price",
                metric_value=min_price,
                supplier_id=prices_sorted[0][0].supplier_id,
                supplier_name=prices_sorted[0][0].supplier.company_name,
            )
        )
        metrics.append(
            QuoteComparisonMetric(
                metric_name="max_price",
                metric_value=max_price,
                supplier_id=prices_sorted[-1][0].supplier_id,
                supplier_name=prices_sorted[-1][0].supplier.company_name,
            )
        )
        metrics.append(
            QuoteComparisonMetric(metric_name="avg_price", metric_value=avg_price)
        )

    elif comparison_type == "delivery":
        # Teslimat süresi karşılaştırması
        deliveries = [
            (sq, int(sq.delivery_time) if sq.delivery_time else 999)
            for sq in supplier_quotes
        ]
        deliveries_sorted = sorted(deliveries, key=lambda x: x[1])

        if deliveries_sorted[0][1] != 999:
            metrics.append(
                QuoteComparisonMetric(
                    metric_name="fastest_delivery",
                    metric_value=deliveries_sorted[0][1],
                    supplier_id=deliveries_sorted[0][0].supplier_id,
                    supplier_name=deliveries_sorted[0][0].supplier.company_name,
                )
            )

    return QuoteComparisonOut(
        quote_id=quote_id,
        comparison_type=comparison_type,
        metrics=metrics,
        generated_at=utcnow(),
    )


# ============ SUPPLIER RATING ============


@router.post("/{quote_id}/rate-supplier", response_model=SupplierRatingOut)
def rate_supplier(
    quote_id: int,
    rating_data: SupplierRatingCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Tedarikçiye puan ver"""
    _get_scoped_quote_or_404(db, quote_id, current_user)

    # Tedarikçi kontrol et
    supplier = db.query(Supplier).filter(Supplier.id == rating_data.supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Tedarikçi bulunamadı")

    # Genel puanı hesapla (ortalama)
    overall_rating = (
        rating_data.price_rating
        + rating_data.delivery_rating
        + rating_data.quality_rating
        + rating_data.communication_rating
    ) / 4

    # Mevcut puanı güncelle veya oluştur
    existing_rating = (
        db.query(SupplierRating)
        .filter(
            SupplierRating.quote_id == quote_id,
            SupplierRating.supplier_id == rating_data.supplier_id,
            SupplierRating.rated_by_id == current_user.id,
        )
        .first()
    )

    if existing_rating:
        existing_rating.price_rating = rating_data.price_rating
        existing_rating.delivery_rating = rating_data.delivery_rating
        existing_rating.quality_rating = rating_data.quality_rating
        existing_rating.communication_rating = rating_data.communication_rating
        existing_rating.overall_rating = overall_rating
        existing_rating.comment = rating_data.comment
        rating = existing_rating
    else:
        rating = SupplierRating(
            supplier_id=rating_data.supplier_id,
            quote_id=quote_id,
            rated_by_id=current_user.id,
            price_rating=rating_data.price_rating,
            delivery_rating=rating_data.delivery_rating,
            quality_rating=rating_data.quality_rating,
            communication_rating=rating_data.communication_rating,
            overall_rating=overall_rating,
            comment=rating_data.comment,
        )
        db.add(rating)

    db.commit()
    db.refresh(rating)
    return rating


@router.get("/supplier/{supplier_id}/ratings", response_model=list[SupplierRatingOut])
def get_supplier_ratings(
    supplier_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Tedarikçinin tüm puanlarını al"""
    ratings = (
        db.query(SupplierRating).filter(SupplierRating.supplier_id == supplier_id).all()
    )
    return ratings


# ============ REPORTING DASHBOARD ============


@router.get("/{quote_id}/dashboard", response_model=ReportingDashboardData)
def get_reporting_dashboard(
    quote_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Raporlama dashboard'u - tüm veriler"""
    quote = _get_scoped_quote_or_404(db, quote_id, current_user)

    # Tedarikçi sayısı
    supplier_quotes = (
        db.query(SupplierQuote).filter(SupplierQuote.quote_id == quote_id).all()
    )

    submitted_count = len([sq for sq in supplier_quotes if sq.submitted_at])

    # Fiyat analizi
    try:
        price_analysis_obj = (
            db.query(PriceAnalysis).filter(PriceAnalysis.quote_id == quote_id).first()
        )
    except Exception:
        price_analysis_obj = None

    # Puanlar
    ratings = db.query(SupplierRating).filter(SupplierRating.quote_id == quote_id).all()

    avg_rating = (
        sum(r.overall_rating for r in ratings) / len(ratings) if ratings else None
    )

    # Sözleşmeler
    contracts = db.query(Contract).filter(Contract.quote_id == quote_id).all()

    return ReportingDashboardData(
        quote_id=quote_id,
        quote_title=quote.title,
        price_analysis=price_analysis_obj,
        supplier_ratings=[SupplierRatingOut.from_orm(r) for r in ratings],
        total_suppliers=len(supplier_quotes),
        submitted_responses=submitted_count,
        average_rating=avg_rating,
        contracts=[ContractOut.from_orm(c) for c in contracts],
    )


@router.get("/{quote_id}/comparison/export-xlsx")
def export_quote_comparison_xlsx(
    quote_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Karşılaştırma verisini ikinci görünüm formatına yakın detaylı Excel olarak dışa aktarır."""
    _get_scoped_quote_or_404(db, quote_id, current_user)

    dataset = _build_comparison_dataset(db, quote_id)

    suppliers = dataset["suppliers"]
    items = dataset["items"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Karsilastirma Raporu"

    ws.append(["Teklif", dataset["quote"]["title"]])
    ws.append(["Teklif ID", dataset["quote"]["id"]])
    ws.append(["Rapor Tarihi", utcnow().strftime("%Y-%m-%d %H:%M")])
    ws.append([])
    ws.append(
        [
            "Tedarikci",
            "Revizyon",
            "Toplam Tutar",
            "Indirim Tutar",
            "Final Tutar",
            "Teslimat (Gun)",
            "Durum",
            "Onay",
        ]
    )

    for row in suppliers:
        approved = bool(row["approved"])
        ws.append(
            [
                row["supplier_name"],
                row["revision_number"],
                row["total_amount"],
                row["discount_amount"],
                row["final_amount"],
                row["delivery_time"],
                row["status"],
                "EVET" if approved else "",
            ]
        )

    ws.append([])
    ws.append(
        [
            "Onaylanan Tedarikci",
            dataset["quote"]["approved_supplier_name"] or "Henüz yok",
        ]
    )

    for col in ("A", "B", "C", "D", "E", "F", "G", "H"):
        ws.column_dimensions[col].width = 20

    for row_idx in range(6, 6 + len(suppliers)):
        ws[f"C{row_idx}"].number_format = "#,##0.00"
        ws[f"D{row_idx}"].number_format = "#,##0.00"
        ws[f"E{row_idx}"].number_format = "#,##0.00"

    # Kalem bazli detayli karsilastirma bolumu
    ws.append([])
    detail_header_row = ws.max_row + 1
    ws.cell(detail_header_row, 1, "Sira")
    ws.cell(detail_header_row, 2, "Aciklama")
    ws.cell(detail_header_row, 3, "Birim")
    ws.cell(detail_header_row, 4, "Miktar")
    ws.cell(detail_header_row, 5, "Tahmini Birim Fiyat")

    col = 6
    for supplier in suppliers:
        ws.merge_cells(
            start_row=detail_header_row,
            start_column=col,
            end_row=detail_header_row,
            end_column=col + 1,
        )
        ws.cell(detail_header_row, col, supplier["supplier_name"])
        ws.cell(detail_header_row + 1, col, "Birim Fiyat")
        ws.cell(detail_header_row + 1, col + 1, "Birim Toplam")
        col += 2

    header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    sub_header_fill = PatternFill(fill_type="solid", fgColor="EEF3FA")
    group_fill = PatternFill(fill_type="solid", fgColor="FDE9A9")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(1, col):
        ws.cell(detail_header_row, c).fill = header_fill
        ws.cell(detail_header_row, c).font = Font(bold=True)
        ws.cell(detail_header_row, c).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(detail_header_row, c).border = border

        ws.cell(detail_header_row + 1, c).fill = sub_header_fill
        ws.cell(detail_header_row + 1, c).font = Font(bold=True)
        ws.cell(detail_header_row + 1, c).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(detail_header_row + 1, c).border = border

    current_row = detail_header_row + 2
    for item in items:
        ws.cell(current_row, 1, item["line_number"])
        desc = item["description"]
        if item["detail"]:
            desc = f"{desc}\n{item['detail']}"
        ws.cell(current_row, 2, desc)
        ws.cell(current_row, 3, item["unit"])
        ws.cell(current_row, 4, item["quantity"])
        ws.cell(current_row, 5, item["base_unit_price"])

        price_col = 6
        for supplier in suppliers:
            supplier_price = (
                item["supplier_prices"].get(str(supplier["supplier_quote_id"])) or {}
            )
            ws.cell(current_row, price_col, supplier_price.get("unit_price"))
            ws.cell(current_row, price_col + 1, supplier_price.get("total_price"))
            price_col += 2

        for c in range(1, col):
            cell = ws.cell(current_row, c)
            cell.border = border
            if c in (2,):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif c in (1, 3):
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")

            if c >= 5:
                cell.number_format = "#,##0.00"

            if item["is_group_header"]:
                cell.fill = group_fill
                cell.font = Font(bold=True)

        if item["is_group_header"]:
            for c in range(3, col):
                ws.cell(current_row, c, None)

        current_row += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 16

    dynamic_col = 6
    for _ in suppliers:
        ws.column_dimensions[openpyxl.utils.get_column_letter(dynamic_col)].width = 14
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(dynamic_col + 1)
        ].width = 14
        dynamic_col += 2

    chart = BarChart()
    chart.title = "Tedarikci Final Tutar Karsilastirmasi"
    chart.y_axis.title = "Final Tutar"
    chart.x_axis.title = "Tedarikci"
    data = Reference(ws, min_col=5, min_row=5, max_row=5 + len(suppliers))
    cats = Reference(ws, min_col=1, min_row=6, max_row=5 + len(suppliers))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 18
    ws.add_chart(chart, "J5")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"quote_{quote_id}_karsilastirma_raporu.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/{quote_id}/comparison/detailed")
def get_quote_comparison_detailed(
    quote_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Karşılaştırma ekranı ve Excel için son revizyon bazlı detaylı veri döner."""
    _get_scoped_quote_or_404(db, quote_id, current_user)
    dataset = _build_comparison_dataset(db, quote_id)
    return dataset
